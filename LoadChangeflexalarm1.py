# -*- coding: utf-8 -*-
import os
import io
from datetime import datetime
from dataclasses import dataclass
from typing import List, Optional

from PySide6.QtCore import Qt, QTimer
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QComboBox, QMessageBox, QFrame, QTabWidget
)

# Matplotlib for Qt
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

# Excel + Audio
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from gtts import gTTS
import pygame

# === Import mô-đun tính toán của bạn ===
# Giữ nguyên như bạn đã refactor:
from modules.power_logic import CalcConfig, compute_power_change_and_pauses


# =========================
# ExcelUpdater (giữ nguyên)
# =========================
class ExcelUpdater:
    """CLASS XUẤT DỮ LIỆU TĂNG GIẢM TẢI TRONG CA"""
    def __init__(self, file_name: str):
        self.file_name = file_name
        if os.path.exists(self.file_name):
            try:
                self.wb = load_workbook(filename=self.file_name)
            except InvalidFileException:
                self.wb = Workbook()
        else:
            self.wb = Workbook()
        self.sheet = self.wb.active

    def append_data(self, data: dict):
        row = self.sheet.max_row + 1
        col = 1
        for _, value in data.items():
            column_letter = get_column_letter(col)
            self.sheet[f'{column_letter}{row}'] = value
            col += 1
        self.wb.save(self.file_name)

    def append_data1(self, data1: dict):
        row = self.sheet.max_row + 1
        col = 1
        for _, value in data1.items():
            column_letter = get_column_letter(col)
            self.sheet[f'{column_letter}{row}'] = value
            col += 1
        self.wb.save(self.file_name)


# =====================================
# Tiện ích âm thanh (gTTS + pygame)
# =====================================
def tts_and_play(message: str):
    try:
        tts = gTTS(text=message, lang="en")
        audio_stream = io.BytesIO()
        tts.write_to_fp(audio_stream)
        audio_stream.seek(0)
        if not pygame.mixer.get_init():
            pygame.mixer.init()
        pygame.mixer.music.load(audio_stream, "mp3")
        pygame.mixer.music.play()
    except Exception as e:
        print(f"❌ Lỗi phát âm thanh: {e}")


# =====================================
# Widget chính PySide6
# =====================================
class PowerChangeWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)

        # --- trạng thái ---
        self.excel_updater = ExcelUpdater('abc.xlsx')
        self.alarm_played_for_429 = False
        self.alarm_played_for_post_pause = False
        self.alarm_played_for_final_load = False
        self.alarm_played_for_hold_complete = False
        self.messagebox_shown = False

        # thresholds + alarm texts
        self.threshold_429 = 429.0
        self.holding_complete_mw = 462.0
        self.final_load_mw = 560.0

        self.alarm_texts = {
            "429": "Reached 429 MW",
            "holding_complete": "Holding complete at 462 MW",
            "final_load": "Final target load achieved, To pay Attention to check control mode and SCC Speed mode",
            "hold_10_min": "Holding 10 minutes completed",
        }

        # users’ pause config + pulverizer mode
        self.pause_time_429_min = 0
        self.pause_time_hold_min = 30
        self.pulverizer_mode_default = "3 Puls"  # default

        # data series for plotting
        self.times1: List[datetime] = []
        self.powers1: List[float] = []
        self.times2: List[datetime] = []   # reserved (hidden layout 2)
        self.powers2: List[float] = []

        # result times
        self.final_load_time: Optional[datetime] = None
        self.time_reaching_429: Optional[datetime] = None
        self.post_pause_time: Optional[datetime] = None
        self.time_holding_462: Optional[datetime] = None
        self.hold_complete_time: Optional[datetime] = None

        # --- UI ---
        self._build_ui()

        # Timer kiểm tra mốc báo động (không hiển thị đồng hồ)
        self.check_timer = QTimer(self)
        self.check_timer.timeout.connect(self.check_and_alarm)
        self.check_timer.start(1000)

    # ----------------------
    # UI construction
    # ----------------------
    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(10, 10, 10, 10)
        root.setSpacing(8)

        # Dòng nhãn kết quả
        result_row = QHBoxLayout()
        self.total_load_time_label = QLabel("Load reaching: ")
        self.time_reaching_429_label = QLabel("429 MW: ")
        self.post_pause_time_label = QLabel("The holding comp: ")
        self.hold_complete_label = QLabel("Holding 10M:")

        for w in [
            self.total_load_time_label, self.time_reaching_429_label,
            self.post_pause_time_label, self.hold_complete_label
        ]:
            w.setStyleSheet("font-weight:600;")
            result_row.addWidget(w, 0, Qt.AlignLeft)

        result_row.addStretch(1)
        root.addLayout(result_row)

        # Khung nhập chính
        input_row = QHBoxLayout()
        self.start_power_edit = self._labeled_edit(input_row, "Start Power (MW):")
        self.target_power_edit = self._labeled_edit(input_row, "Target Power (MW):")
        self.start_time_edit = self._labeled_edit(input_row, "Start Time (HH:MM):")

        self.enter_btn = QPushButton("Enter")
        self.enter_btn.clicked.connect(self.on_enter_clicked)
        self.reset_btn = QPushButton("Reset")
        self.reset_btn.clicked.connect(self.on_reset_clicked)

        # nút gear toggle layout ẩn
        self.toggle_btn = QPushButton("⚙")
        self.toggle_btn.clicked.connect(self.toggle_hidden_layout)

        input_row.addWidget(self.enter_btn)
        input_row.addWidget(self.reset_btn)
        input_row.addWidget(self.toggle_btn)
        root.addLayout(input_row)

        # Khung nhập bổ sung (Hold)
        hold_row = QHBoxLayout()
        self.holding_load_edit = self._labeled_edit(hold_row, "Holding Load (MW):")
        self.holding_time_edit = self._labeled_edit(hold_row, "Holding Time (HH:MM):")
        self.hold_btn = QPushButton("Hold")
        self.hold_btn.clicked.connect(self.on_hold_clicked)
        hold_row.addWidget(self.hold_btn)
        root.addLayout(hold_row)

        # Matplotlib Figure
        self.figure = Figure(figsize=(6, 4), dpi=120)
        self.ax = self.figure.add_subplot(111)
        self.ax.set_title('TREND: POWER DEPEND ON TIMES')
        self.ax.set_xlabel('TIMES')
        self.ax.set_ylabel('POWER (MW)')
        self.canvas = FigureCanvas(self.figure)
        root.addWidget(self.canvas, 1)

        # Layout ẩn (tuỳ biến cảnh báo + pause + pulverizer)
        self.hidden_frame = QFrame()
        self.hidden_frame.setFrameShape(QFrame.StyledPanel)
        hidden_layout = QVBoxLayout(self.hidden_frame)

        title = QLabel("Custom Alarm Messages")
        title.setStyleSheet("font-weight:700;")
        hidden_layout.addWidget(title)

        self.edit_alarm_429 = self._labeled_edit(hidden_layout, "429 MW Alert:", default=self.alarm_texts["429"])
        self.edit_alarm_hold_comp = self._labeled_edit(hidden_layout, "Holding Complete Alert:", default=self.alarm_texts["holding_complete"])
        self.edit_alarm_final = self._labeled_edit(hidden_layout, "Final Load Alert:", default=self.alarm_texts["final_load"])
        self.edit_alarm_hold10 = self._labeled_edit(hidden_layout, "Hold 10 Min Alert:", default=self.alarm_texts["hold_10_min"])

        # pause + pulverizer
        pause_row = QHBoxLayout()
        self.pause_429_edit = self._labeled_edit(pause_row, "Pause at 429 MW (min):", default=str(self.pause_time_429_min), width=60)
        self.pause_hold_edit = self._labeled_edit(pause_row, "Hold duration at holding MW (min):", default=str(self.pause_time_hold_min), width=60)

        label = QLabel("Pulverizer Mode:")
        self.pulverizer_combo = QComboBox()
        self.pulverizer_combo.addItems(["3 Puls", "4 Puls"])
        self.pulverizer_combo.setCurrentText(self.pulverizer_mode_default)
        pause_row.addWidget(label)
        pause_row.addWidget(self.pulverizer_combo)
        hidden_layout.addLayout(pause_row)

        # Ẩn ngay từ đầu
        self.hidden_frame.setVisible(False)
        root.addWidget(self.hidden_frame)

    def _labeled_edit(self, parent_layout_or_widget, label: str, default: str = "", width: int = 100) -> QLineEdit:
        if isinstance(parent_layout_or_widget, QVBoxLayout):
            row = QHBoxLayout()
            parent_layout_or_widget.addLayout(row)
        else:
            row = parent_layout_or_widget
        lab = QLabel(label)
        edit = QLineEdit()
        if default:
            edit.setText(default)
        if width:
            edit.setFixedWidth(width)
        row.addWidget(lab)
        row.addWidget(edit)
        return edit

    # ----------------------
    # Event handlers
    # ----------------------
    def toggle_hidden_layout(self):
        self.hidden_frame.setVisible(not self.hidden_frame.isVisible())

    def on_enter_clicked(self):
        try:
            start_power = float(self.start_power_edit.text())
            target_power = float(self.target_power_edit.text())
            start_time_str = self.start_time_edit.text().strip()
        except ValueError:
            QMessageBox.critical(self, "Input Error", "Vui lòng nhập đúng định dạng số/giờ.")
            return

        # cập nhật cấu hình từ UI ẩn (nếu người dùng đã mở và chỉnh)
        self.alarm_texts["429"] = self.edit_alarm_429.text()
        self.alarm_texts["holding_complete"] = self.edit_alarm_hold_comp.text()
        self.alarm_texts["final_load"] = self.edit_alarm_final.text()
        self.alarm_texts["hold_10_min"] = self.edit_alarm_hold10.text()

        try:
            self.pause_time_429_min = int(self.pause_429_edit.text() or "0")
            self.pause_time_hold_min = int(self.pause_hold_edit.text() or "30")
        except ValueError:
            QMessageBox.critical(self, "Input Error", "Thời gian pause/hold phải là số phút.")
            return

        pulverizer_mode = self.pulverizer_combo.currentText()

        # Tính toán bằng modules.power_logic
        try:
            cfg = CalcConfig(
                threshold_429=self.threshold_429,
                hold_power=self.holding_complete_mw,
                pause_time_429_min=self.pause_time_429_min,
                pause_time_hold_min=self.pause_time_hold_min,
                pulverizer_mode=pulverizer_mode,
            )
        except Exception as e:
            QMessageBox.critical(self, "Config Error", f"Lỗi cấu hình: {e}")
            return

        result = compute_power_change_and_pauses(
            start_power=start_power,
            target_power=target_power,
            start_time=start_time_str,  # "HH:MM" hoặc datetime
            cfg=cfg,
        )

        # Gán cho vẽ/logic
        self.times1 = result.times
        self.powers1 = result.powers
        self.final_load_time = result.final_load_time
        self.time_reaching_429 = result.time_reaching_429
        self.post_pause_time = result.post_pause_time
        self.time_holding_462 = result.time_holding_462
        self.hold_complete_time = result.hold_complete_time

        # Cập nhật nhãn
        self.total_load_time_label.setText(
            "Load reaching: " + (self.final_load_time.strftime("%H:%M") if self.final_load_time else "")
        )
        if self.time_reaching_429:
            self.time_reaching_429_label.setText(f"429 MW: {self.time_reaching_429.strftime('%H:%M')}")
            self.post_pause_time_label.setText(
                "The holding complete: " + (self.post_pause_time.strftime("%H:%M") if self.post_pause_time else "")
            )
        else:
            self.time_reaching_429_label.setText("Do not reach 429 MW in the load changing process.")
            self.post_pause_time_label.setText("")

        if self.hold_complete_time:
            self.hold_complete_label.setText(f"Holding {self.pause_time_hold_min}M: {self.hold_complete_time.strftime('%H:%M')}")
        else:
            self.hold_complete_label.setText("Holding 10M:")

        # Vẽ đồ thị
        self.update_plot()

        # Ghi log Excel
        copy_text = (
            f"Decrease Unit load to {target_power} MW/Giảm tải xuống {target_power} MW"
            if start_power > target_power else
            f"Increase Unit load to {target_power} MW/Tăng tải lên {target_power} MW"
        )
        data = {
            'time_now': datetime.now(),
            'start_power': start_power,
            'target_power': target_power,
            'start_time_str': start_time_str,
            'copy_text': copy_text,
        }
        self.excel_updater.append_data(data)

    def on_hold_clicked(self):
        try:
            holding_load = float(self.holding_load_edit.text())
        except ValueError:
            QMessageBox.critical(self, "Input Error", "Holding Load (MW) phải là số.")
            return

        holding_time = self.holding_time_edit.text().strip()
        QMessageBox.information(self, "Holding Information", f"Holding Load: {holding_load}\nHolding Time: {holding_time}")

        data1 = {
            'time_now_hold': datetime.now(),
            '........': "",
            '.........': "",
            'holding_time': holding_time,
            'holding_load': f"Hold the load at {holding_load} MW/ Giữ tải tại {holding_load} MW",
        }
        self.excel_updater.append_data1(data1)

    def on_reset_clicked(self):
        # Xoá inputs
        self.start_power_edit.clear()
        self.target_power_edit.clear()
        self.start_time_edit.clear()
        self.holding_load_edit.clear()
        self.holding_time_edit.clear()

        # Reset cờ báo động
        self.alarm_played_for_429 = False
        self.alarm_played_for_post_pause = False
        self.alarm_played_for_final_load = False
        self.alarm_played_for_hold_complete = False
        self.messagebox_shown = False

        # Reset nhãn kết quả
        self.total_load_time_label.setText("Load reaching: ")
        self.time_reaching_429_label.setText("429 MW: ")
        self.post_pause_time_label.setText("The holding comp: ")
        self.hold_complete_label.setText("Holding 10M:")

        # Xoá series + kết quả thời gian
        self.times1.clear()
        self.powers1.clear()
        self.times2.clear()
        self.powers2.clear()

        self.final_load_time = None
        self.time_reaching_429 = None
        self.post_pause_time = None
        self.time_holding_462 = None
        self.hold_complete_time = None

        # Làm mới đồ thị
        self.ax.clear()
        self.ax.set_title('TREND: POWER DEPEND ON TIMES')
        self.ax.set_xlabel('TIMES')
        self.ax.set_ylabel('POWER (MW)')
        self.canvas.draw()

    # ----------------------
    # Plotting
    # ----------------------
    def update_plot(self):
        self.ax.clear()

        # Vẽ main series nếu có
        if self.times1 and self.powers1:
            self.ax.plot(self.times1, self.powers1, marker='o', linestyle='-', label='Main Load Change')

        # Vẽ series thứ 2 (nếu sau này bạn dùng)
        if self.times2 and self.powers2:
            self.ax.plot(self.times2, self.powers2, marker='o', linestyle='-', label='Hidden Layout 2')

        if self.times1 or self.times2:
            self.ax.legend()

        self.ax.set_title('TREND: POWER DEPEND ON TIMES')
        self.ax.set_xlabel('TIMES')
        self.ax.set_ylabel('POWER (MW)')
        self.figure.autofmt_xdate()
        self.canvas.draw()

    # ----------------------
    # Alarm checking (no clock UI)
    # ----------------------
    def check_and_alarm(self):
        now = datetime.now()
        s_now = now.hour * 3600 + now.minute * 60 + now.second

        def due(dt: Optional[datetime]) -> bool:
            if not dt:
                return False
            s_alarm = dt.hour * 3600 + dt.minute * 60
            return s_now >= s_alarm

        # Map mốc -> (thời điểm, cờ đã báo)
        alarm_map = [
            ("429", self.time_reaching_429, "alarm_played_for_429"),
            ("holding_complete", self.post_pause_time, "alarm_played_for_post_pause"),
            ("final_load", self.final_load_time, "alarm_played_for_final_load"),
            ("hold_10_min", self.hold_complete_time, "alarm_played_for_hold_complete"),
        ]

        for key, t_alarm, flag_name in alarm_map:
            if t_alarm and not getattr(self, flag_name):
                if due(t_alarm):
                    tts_and_play(self.alarm_texts[key])
                    setattr(self, flag_name, True)

        # Khi đạt final_load -> popup nhắc chế độ
        if not self.messagebox_shown and self.alarm_played_for_final_load:
            try:
                target_power_value = float(self.target_power_edit.text())
                if target_power_value >= self.final_load_mw:
                    control_mode_message = "Check the control mode: LL MODE"
                    scc_mode_message = "Check the SCC mode: SCC HIGHT MODE"
                else:
                    control_mode_message = "Check the control mode: GOV MODE"
                    scc_mode_message = "Check the SCC mode: SCC AUTO MODE"

                QMessageBox.information(self, "Control Mode", control_mode_message)
                QMessageBox.information(self, "SCC Mode", scc_mode_message)
                self.messagebox_shown = True
            except ValueError:
                pass


# =========================
# App launcher
# =========================
def main():
    import sys
    app = QApplication(sys.argv)

    win = QWidget()
    win.setWindowTitle("Power Change Application (PySide6)")
    win.resize(1000, 800)

    tabs = QTabWidget()
    tab1 = PowerChangeWidget()
    tabs.addTab(tab1, "Power Change")

    layout = QVBoxLayout(win)
    layout.addWidget(tabs)
    win.setLayout(layout)
    win.show()

    sys.exit(app.exec())


if __name__ == "__main__":
    main()

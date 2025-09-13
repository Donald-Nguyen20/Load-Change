# modules/excel_io.py
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

class ExcelUpdater:  # CLASS XUẤT DỮ LIỆU TĂNG GIẢM TẢI TRONG CA
    def __init__(self, file_name):
        self.file_name = file_name
        if os.path.exists(self.file_name):
            try:
                self.wb = load_workbook(filename=self.file_name)
            except InvalidFileException:
                self.wb = Workbook()
        else:
            self.wb = Workbook()
        self.sheet = self.wb.active

    def append_data(self, data):
        """Thêm hàng dữ liệu từ một dictionary vào Excel.
        Giả sử 'data' là một dictionary chứa các giá trị cần thêm.
        """
        row = self.sheet.max_row + 1  # Tìm hàng trống đầu tiên
        col = 1  # Bắt đầu từ cột đầu tiên
        for _, value in data.items():
            column_letter = get_column_letter(col)
            self.sheet[f'{column_letter}{row}'] = value
            col += 1
        self.wb.save(self.file_name)

    def append_data1(self, data1):
        """Thêm hàng dữ liệu từ một dictionary vào Excel.
        Giả sử 'data' là một dictionary chứa các giá trị cần thêm.
        """
        row = self.sheet.max_row + 1
        col = 1
        for _, value in data1.items():
            column_letter = get_column_letter(col)
            self.sheet[f'{column_letter}{row}'] = value
            col += 1
        self.wb.save(self.file_name)
